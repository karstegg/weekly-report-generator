#!/usr/bin/env python3
"""
Find equipment availability percentages in Excel files
"""

import openpyxl
from pathlib import Path

def scan_for_availability(file_path: str, shaft_name: str):
    """Scan entire workbook for availability data"""
    print(f"\n{'='*80}")
    print(f"SCANNING: {shaft_name}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    equipment_types = ['DT', 'FL', 'HD', 'RT', 'SR']

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")

        # Look for rows that contain "availability" or equipment type headers
        for row_idx in range(1, min(ws.max_row + 1, 200)):
            row_values = []
            for col_idx in range(1, min(ws.max_column + 1, 30)):
                val = ws.cell(row_idx, col_idx).value
                row_values.append(val)

            # Convert row to string for searching
            row_text = ' '.join([str(v).lower() if v is not None else '' for v in row_values])

            # Check if this row contains availability information
            if 'availab' in row_text or 'avail %' in row_text:
                # Print this row and the next 10 rows
                print(f"\nFound 'availability' at row {row_idx}:")
                for i in range(max(1, row_idx - 2), min(ws.max_row + 1, row_idx + 15)):
                    display_row = []
                    for col_idx in range(1, min(ws.max_column + 1, 20)):
                        val = ws.cell(i, col_idx).value
                        if val is not None:
                            if isinstance(val, (int, float)):
                                if 0 < val < 1:
                                    display_row.append(f"{val:.4f}")
                                elif 0 <= val <= 100:
                                    display_row.append(f"{val:.2f}")
                                else:
                                    display_row.append(str(val))
                            else:
                                display_row.append(str(val)[:30])
                        else:
                            display_row.append('')

                    if any(display_row):
                        print(f"  Row {i}: {' | '.join(display_row[:15])}")

                print()  # Blank line after section
                break  # Move to next sheet after finding availability

            # Also check for equipment type names in headers
            if any(equip.lower() in row_text for equip in ['dump truck', 'loader', 'drill', 'bolter', 'scaler']):
                # Check if next few rows have percentage values
                has_percentages = False
                for check_row in range(row_idx, min(ws.max_row + 1, row_idx + 20)):
                    for col_idx in range(1, min(ws.max_column + 1, 20)):
                        val = ws.cell(check_row, col_idx).value
                        if isinstance(val, (int, float)) and 0 < val <= 1:
                            has_percentages = True
                            break
                    if has_percentages:
                        break

                if has_percentages:
                    print(f"\nFound equipment types at row {row_idx} with percentage data:")
                    for i in range(row_idx, min(ws.max_row + 1, row_idx + 15)):
                        display_row = []
                        for col_idx in range(1, min(ws.max_column + 1, 15)):
                            val = ws.cell(i, col_idx).value
                            if val is not None:
                                if isinstance(val, (int, float)):
                                    if 0 < val < 1:
                                        display_row.append(f"{val:.4f}")
                                    else:
                                        display_row.append(str(val)[:20])
                                else:
                                    display_row.append(str(val)[:30])
                            else:
                                display_row.append('')

                        if any(display_row):
                            print(f"  Row {i}: {' | '.join(display_row)}")
                    print()
                    break

def main():
    base_path = Path("/home/user/weekly-report-generator/data/week33")

    scan_for_availability(
        str(base_path / "N3 Eng Report Week  06 Feb - 12 Feb 2026.xlsx"),
        "N3"
    )

    scan_for_availability(
        str(base_path / "Nch2 Weekly Report 13 February 2026.xlsx"),
        "Nch2"
    )

if __name__ == "__main__":
    main()
